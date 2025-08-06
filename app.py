'''
Verificar se valores estao pagos ou nao e caso estejam, verificar a forma 
de pagamento.
com base em uma planilha, pegar o cpf e consultar no sitedisponibilazado 
para verificar se a conta esta (ou nao) paga
caso esteja paga, preencher na planilha de fechamento como "ok" 
caso contrario, informar que continua pendente.

--Definir quais sao os amnuais dessa tarefa, para que eu possa transforma:
1- Entrar na planilha e extrair o cpf do cliente
2- Entro no site e uso cpf da planilha para pesquisar o status do pagamento daquele cliente
3- Verificar se esta em dia ou atrasado
4- se estevir em dia pegar a data do pagamento e o metodo de pagamento
5- caso contrario se estiver atrasado colocar o status como pendente.
6- inserir informaçao (nome, valor, cpf e vencimento e status e caso em dia, datapagamento e a coluna metodo pagamento (cartao ou boleto))em uma nova planilha.
7- Repetiur ate chegar no ultimo.
'''
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep


planilha_clientes = openpyxl.load_workbook('dados_ficticios.xlsx')
pagina_clientes = planilha_clientes['in']
driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

sleep(5)  

campo_pesqusia = driver.find_element(
    By.XPATH, "//input[@id='cpfInput']")  
sleep(1) 
campo_pesqusia.clear()  
campo_pesqusia.send_keys(cpf)
sleep(1)  


botao_pesquisar = driver.find_element(
    By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
botao_pesquisar.click()
sleep(4) 

status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
if status.text == "em dia":
    data_pagamento = driver.find_element(
        By.XPATH, "//p[@id='paymentDate']").text
    metodo_pagamento = driver.find_element(
        By.XPATH, "//p[@id='paymentMethod']").text

    data_pagamento_limpo = data_pagamento.split()[3]

    metodo_pagamento_limpo = metodo_pagamento.split()[3]

    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']
    pagina_fechamento.append([nome, valor, cpf, vencimento,
                             'em dia', data_pagamento_limpo, metodo_pagamento_limpo])

    planilha_fechamento.save('planilha fechamento.xlsx')

else:
    planilha_fechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    pagina_fechamento = planilha_fechamento['Sheet1']

    pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente'])
    planilha_fechamento.save('planilha fechamento.xlsx')
